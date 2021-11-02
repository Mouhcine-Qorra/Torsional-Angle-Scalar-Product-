import random
import time
from openpyxl import load_workbook
from send_mail import email_sender_func


wb = load_workbook('test.xlsx')
ws = wb.active
# sheet1 = wb['Feuil1']
# print(sheet1.cell(2, 2).value)
# print(sheet1.max_column)
# for i in range(1, sheet1.max_row + 1):
#     for j in range(1, sheet1.max_column):
#         print(sheet1.cell(row=i, column=j).value)
# sheet1.cell(row=i, column=j, value= 'anything')
# wb.save('test.xlsx')

all_emails = []
usernames = []
result = []
duplicated = []
emails = []

def find_duplicated():
    wb_auto = load_workbook('automation.xlsx')
    wss = wb_auto.active
    count = 2
    for elem in all_emails:
        email = set()
        if elem in emails:
            duplicated.append(f'line {count}: {elem}')
            print(f'{elem} is duplicated')
        else:
            email.add(elem)
        count += 1

    if duplicated is not None:
        print(f'found duplicated emails:')
        for i in duplicated:
            print(f'{i}\n')
        print('stopping program')
    else:
        pass

def send_mail_using_excel():
    for i in range(50, 144):
        print(f'Traiting line {i}:', end=" ")
        print([j.value for j in ws[f'{i}']])
    
        receiver = ws[f"A{i}"].value
        username = ws[f'B{i}'].value
        herrfrau = ws[f'C{i}'].value

        if username:
            if herrfrau == 'herr':
                einleitung = f'Sehr geehrter Herr {username}'
            elif herrfrau == 'frau':
                einleitung = f'Sehr geehrte Frau {username}'
            else:
                einleitung = 'Sehr geehrte Damen und Herren'
        else:
            einleitung = 'Sehr geehrte Damen und Herren'
    
        c = email_sender_func(receiver, einleitung, i)
        if c:
            result.append(c)
        time_sec = random.randint(15, 30)
        print(f'waiting {time_sec}sec\n')
        time.sleep(time_sec)

    if result:
        print(f'Result:')
        for res in result:
            print(f'{res}\n')
    else:
        print('All emails has been sent Successfully')
send_mail_using_excel()
