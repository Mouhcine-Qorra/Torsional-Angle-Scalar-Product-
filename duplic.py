from openpyxl import load_workbook


emails = {}
duplicated_emails = {}

def dup_emails(files):
    for file in files:
        if 'test' in file:
            c = 2
        else:
            c = 153

        found = False
        wb = load_workbook(file)
        ws = wb.active
        sheet1 = wb['Feuil1']
        for i in range(2, sheet1.max_row + 1):
            if ws[f"A{i}"].value in emails.values():
                found = True
                duplicated_emails[f'{c}'] = ws[f"A{i}"].value
            else:
                emails[f'{c}'] = ws[f"A{i}"].value
            c += 1

        if found:
            print(f'\nDuplicated Emails in {file.replace(".xlsx", "").title()} File:')
            for k, v in duplicated_emails.items():
                if int(k) > 152: k = int(k) - 151
                print(f'{k}:\t{v}')
                for key, value in emails.items():
                    if v == value:
                        print(f'\tin {key}')
                        break
        else:
            print(f'\nNo Duplicated Emails Founded in {file.replace(".xlsx", "").title()} File\n')

files = ['test.xlsx', 'automation.xlsx']
dup_emails(files)
