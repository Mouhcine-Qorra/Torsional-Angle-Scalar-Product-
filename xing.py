# excel_number
from selenium import webdriver
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook



emails = {}
duplicated_emails = {}

def dup_emails(files):
    for file in files:
        if 'test' in file:
            c = 2
        else:
            c = 153

        found = False
        wb = load_workbook(file)
        ws = wb.active
        sheet1 = wb['Feuil1']
        for i in range(2, sheet1.max_row + 1):
            if ws[f"A{i}"].value in emails.values():
                found = True
                duplicated_emails[f'{c}'] = ws[f"A{i}"].value
            else:
                emails[f'{c}'] = ws[f"A{i}"].value
            c += 1

        if found:
            print(f'\nDuplicated Emails in {file.replace(".xlsx", "").title()} File:')
            for k, v in duplicated_emails.items():
                if int(k) > 152: k = int(k) - 151
                print(f'{k}:\t{v}')
                for key, value in emails.items():
                    if v == value:
                        print(f'\tin {key}')
                        break
        else:
            print(f'\nNo Duplicated Emails Founded in {file.replace(".xlsx", "").title()} File\n')

files = ['test.xlsx', 'automation.xlsx']
dup_emails(files)



options = webdriver.ChromeOptions()
options.add_argument("--user-data-dir=D:\\python-ecom\\cookies")
driver = webdriver.Chrome(options=options, executable_path="C:\\Users\\NAJAT\\AppData\\Local\\chromedriver.exe")
wb = load_workbook('automation.xlsx')
ws = wb.active
sheet1 = wb['Feuil1']
excel_number = 40
page = 1
my = 190
exceptions = []



def mail(stelle, company, url, i, page, my):
    print('Getting the mail')
    write = True
    email = driver.find_element_by_xpath(
        '//*[@id="module-locations"]/div/div[2]/div[2]/div/div[2]/a/div/span').get_attribute('textContent')
    print(f'Found the mail: {email}')
    for k, val in emails.items():
        if val == email:
            write = False
            if int(k) > 152:
                file_name = 'Automation'
                k = int(k) - 151
            else:
                file_name = 'Test'
            print(f'\n\t\t\t!!!We already sent mail to that Email: {k}\t{val} in {file_name}!!!\n')

    if write:
        url = driver.current_url
        sheet1.cell(row=excel_number + i, column=1, value=email)
        sheet1.cell(row=excel_number + i, column=4, value=company)
        sheet1.cell(row=excel_number + i, column=5, value=stelle)
        sheet1.cell(row=excel_number + i, column=6, value=url)
        sheet1.cell(row=excel_number + i, column=7, value=f'ARTICLE:{i}\tPAGE:{page}')

        if email:
            try:
                ansprechpartner = driver.find_element_by_xpath(
                    '//*[@id="module-contacts"]/div/div/div[2]/div/div/a/div[2]/div[1]/h2').get_attribute('textContent')
                sheet1.cell(row=excel_number + i, column=2, value=ansprechpartner)
                print(f'Found Ansprechpartner: {ansprechpartner}')
                # herrfrau()
            except:
                print('Did not Found Ansprechpartner')

        emails[f'{my}'] = email
        my += 1
        wb.save('automation.xlsx')


def get_company_data(stelle, company, url, i, page, my):
    # Getting the Company Page
    driver.switch_to.window(driver.window_handles[1])
    driver.implicitly_wait(10)
    time.sleep(1)
    # Stelle name
    if stelle is None:
        try:
            stelle = driver.find_element_by_xpath('//*[@id="content"]/div/div[1]/div/div[1]/div[1]/div/div[2]/div/div[2]/h1').get_attribute('textContent')
            print(f'Found Stelle Name from get_company_data(): {stelle}')
        except: pass

    # Company name
    if company is None:
        try:
            company = driver.find_element_by_xpath('//*[@id="content"]/div/div[1]/div/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/a/h2').get_attribute('textContent')
            print(f'Found Company Name from get_company_data(): {company}')
        except: pass

    driver.implicitly_wait(10)
    time.sleep(1)
    try:
        failed = False
        driver.find_element_by_xpath('//*[@id="content"]/div/div[1]/div/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/a/h2').click()
        print('Found Company Page')
    except:
        try:
            failed = False
            driver.find_element_by_xpath(
                '//*[@id="content"]/div/div[1]/div/div[1]/div[1]/div/div[2]/div/div[1]/a/div/div/div/img').click()
            print('Found From Except Company Page')

        except:
            try:
                failed = False
                driver.find_element_by_xpath(
                    '//*[@id="content"]/div/div[1]/div/div[6]/div/div/a/div/div/div/img').click()
                print('Found From Except 2 Company Page')

            except:
                try:
                    failed = False
                    driver.find_element_by_xpath(
                        '//*[@id="content"]/div/div[1]/div/div[6]/div/div/div/a/h3').click()
                    print('Found From Except 3 Company Page')

                except:
                    failed = True
                    exceptions.append(f"{i} page {page}:\tFailed to get Company Page")
                    print('Failed to get Company Page')

    driver.implicitly_wait(10)
    time.sleep(1)

    # Getting the mail
    if not failed:
        try: mail(stelle, company, url, i, page, my)
        except Exception as e:
            exceptions.append(f"{i} page {page}:\tError at getting the mail: {e}")
            print(f'Error at getting the mail: {e}')


def load_xing(i, page, my):
    driver.implicitly_wait(10)
    time.sleep(1)
    stelle = None
    desc = None
    company = None
    Go = False
    url = None

    try:
        # Getting Stelle Name
        driver.implicitly_wait(10)
        time.sleep(1)
        try:
            stelle = driver.find_element_by_xpath(f'//*[@id="content"]/div/div/main/div[3]/div[{i}]/a/div[2]/h2/div').get_attribute('textContent')
            print(f'Found Stelle Name')
        except: pass
        # Getting Description
        driver.implicitly_wait(10)
        time.sleep(1)
        try:
            desc = driver.find_element_by_xpath(f'//*[@id="content"]/div/div/main/div[3]/div[{i}]/a/div[2]/div[3]').get_attribute('textContent')
            print(f'Found Desc')
        except: pass
        # Check is this Ausbildung
        if ('ausbild' or 'auszubild') in stelle.lower() or ('ausbild' or 'auszubild') in desc.lower():
            if ('anwendungsentwickl') in stelle.lower() or ('anwendungsentwickl') in desc.lower():
                print("Ausbildung Founded")
                Go = True
        else:
            print("\n\t\t\tAusbildung didn't founded\n")
        if Go:
            # Getting Company Name
            try:
                company = driver.find_element_by_xpath(f'//*[@id="content"]/div/div/main/div[3]/div[{i}]/a/div[2]/div[1]').get_attribute('textContent')
                print(f'Found Company Name: {company}')
            except: pass
            # Getting Articles
            try:
                elem = driver.find_element_by_xpath(
                    f'//*[@id="content"]/div/div/main/div[3]/div[{i}]/a').get_attribute("href")
                driver.execute_script(f"window.open('{elem}', 'new window')")
                print(f'Going to Company Page')

                get_company_data(stelle, company, url, i, page, my)
                driver.close()
                driver.switch_to.window(driver.window_handles[0])
            except:
                try:
                    link = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, f"//*[@id='content']/div/div/main/div[3]/div[{i}]/a/div[2]/h2/div")))
                    ActionChains(driver).key_down(Keys.CONTROL).click(link).key_up(Keys.CONTROL).perform()
                    print(f'found Article {i} from except')

                    get_company_data(stelle, company, url, i, page, my)
                    driver.close()
                    driver.switch_to.window(driver.window_handles[0])
                except Exception as e:
                    exceptions.append(f"{i} page {page}:\tcant click on article {i} bcause: {e}")
                    print(f'cant click on article {i} bcause: {e}')

            driver.implicitly_wait(15)
            time.sleep(1)
        else:
            print(f'Skipped Article{i}')

    except Exception as e:
        exceptions.append(f"{i} page {page}:\tfailed bcus {e}")
        print(f'\nfailed bcus {e}')
    print(f'Finished Article {i}')


driver.get('https://www.xing.com/jobs/search?page=1&paging_context=global_search&keywords=Ausbildung%20Fachinformatiker%20Anwendungsentwicklung')

while page <= 40:
    for i in range(2, 22):
        if i > 2:
            print('\n\n')
        print(f"Page {page}, Article {i}")
        load_xing(i, page, my)
        if i == 21:
            page += 1
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            try:
                driver.implicitly_wait(5)
                time.sleep(1)
                driver.find_element_by_xpath('#content > div > div > main > div.result-list-result-list-container-8d38ca5b > div.result-list-result-list-pagination-13788d01 > div > nav > ol > li:last-child > a > svg').click()
                print(f"\t\t\t__Traiting Page {page}__")
            except:
                driver.implicitly_wait(5)
                time.sleep(1)
                driver.get(f'https://www.xing.com/jobs/search?paging_context=global_search&keywords=Ausbildung%20Fachinformatiker%20Anwendungsentwicklung&page={page}')
                print(f"\t\t\t__Traiting Page {page} from except__")

# Exceptions
if exceptions:
    print('Exceptions:')
    for ex in exceptions:
        print(f"\t{ex}")

