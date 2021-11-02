# excel_number
from selenium import webdriver
import time
from openpyxl import load_workbook



wb = load_workbook('test.xlsx')
ws = wb.active
sheet1 = wb['Feuil1']
# sheet1.cell(row=i, column=1, value=f'emailooo{i}')
# wb.save('test.xlsx')

for i in range(1, sheet1.max_row + 1):
    for j in range(1, sheet1.max_column + 1):
        print(sheet1.cell(row=i, column=j).value)

def ausbildungDe():
    driver = webdriver.Chrome(executable_path="C:\\Users\\NAJAT\\AppData\\Local\\chromedriver.exe")

    driver.get('https://www.ausbildung.de/suche/?utf8=%E2%9C%93&form_main_search[what]=Ausbildung%20fachinformatiker%20anwendungsentwicklung&form_main_search[where]=&t_search_type=root&t_what=Ausbildung%20fachinformatiker%20anwendungsentwicklung&t_where=&form_main_search[show_regular_apprenticeships]=1&form_main_search[show_educational_trainings]=1&form_main_search[show_integrated_degree_programs]=1&form_main_search[show_qualifications]=1&form_main_search[show_inhouse_trainings]=1&form_main_search[show_training_programs]=1&form_main_search[show_educational_trainings_and_regular_apprenticeships]=1&form_main_search[expected_graduation]=&form_main_search[profession_public_id]=&form_main_search[industry_public_id]=&form_main_search[starts_no_earlier_than]=&form_main_search[sort_order]=relevance')
    driver.implicitly_wait(20)
    time.sleep(1)
    window_before = driver.window_handles[0]

    print('Accepted coockies')
    driver.find_element_by_xpath('//*[@id="CybotCookiebotDialogBodyLevelButtonLevelOptinAllowAll"]').click()

    emails = []
    companies = []
    names = []
    to_check = []


    i = 1
    excel_number = 153
    while True:
        driver.implicitly_wait(15)
        time.sleep(10)
        job_name = driver.find_element_by_xpath(f'/html/body/div[3]/main/div/div[1]/div/div[3]/div[2]/div[3]/article[{i}]/a/div[1]/div/h3').get_attribute('textContent')
        company = driver.find_element_by_xpath(f'/html/body/div[3]/main/div/div[1]/div/div[3]/div[2]/div[3]/article[{i}]/a/div[1]/div/h4/strong').get_attribute('textContent')
        driver.find_element_by_xpath(f'/html/body/div[3]/main/div/div[1]/div/div[3]/div[2]/div[3]/article[{i}]/a').click()
        print(f'Accessed to Article {i}')
        window_after = driver.window_handles[1]
        driver.switch_to.window(window_after)
        driver.implicitly_wait(30)
        time.sleep(1)

        try:
            driver.find_element_by_xpath('/html/body/div[2]/main/div/div/div/div[2]/div/div[1]/div[2]')
            print('\tfound Ansprechpartner')
            ansprechpartner = driver.find_element_by_xpath('/html/body/div[2]/main/div/div/div/div[2]/div/div[1]/div[2]/div/div[2]/div[2]/div[1]').get_attribute('textContent')
            print(f'\tname of Ansprechpartner {ansprechpartner}')
            email = driver.find_element_by_xpath('/html/body/div[2]/main/div/div/div/div[2]/div/div[1]/div[2]/div/div[3]/div[1]/div[2]/a').get_attribute('textContent')
            print(f'\temail of Ansprechpartner {email}')
            sheet1.cell(row=excel_number + i, column=1, value=email)
            sheet1.cell(row=excel_number + i, column=2, value=ansprechpartner)
            sheet1.cell(row=excel_number + i, column=4, value=company)
            sheet1.cell(row=excel_number + i, column=5, value=job_name)
            sheet1.cell(row=excel_number + i, column=6, value=i)
            wb.save('test.xlsx')
        except:
            print('\tDid not found Ansprechpartner')
            try:
                try:
                    driver.find_element_by_xpath('/html/body/div[2]/main/header/div/div/div[2]/h2/a').click()
                    driver.implicitly_wait(30)
                    time.sleep(1)
                except:
                    driver.find_element_by_xpath('/html/body/div[2]/main/header/div/div/div[3]/a/img').click()
                    driver.implicitly_wait(30)
                    time.sleep(1)
                try:
                    element = driver.find_element_by_xpath('//*[@id="t-link-application-copy-email"]')
                    webdriver.ActionChains(driver).move_to_element(element).click(element).perform()
                    email = driver.find_element_by_xpath('//*[@id="t-link-application-copy-email"]').text
                    emails.append(email)
                    companies.append(company)
                    names.append(job_name)
                    print(f'\tGot email successfully\n\n')
                    sheet1.cell(row=excel_number + i, column=1, value=email)
                    sheet1.cell(row=excel_number + i, column=4, value=company)
                    sheet1.cell(row=excel_number + i, column=5, value=job_name)
                    sheet1.cell(row=excel_number + i, column=6, value=i)
                    wb.save('test.xlsx')
                except Exception as e:
                    to_check.append(i)
                    print('\tEmail does not exist')
            except:
                print(f'\tError with article {i}\n\n')
                sheet1.cell(row=excel_number + i, column=7, value=i)
                wb.save('test.xlsx')
        driver.close()
        driver.switch_to.window(window_before)

        if i % 18 == 0:
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            driver.implicitly_wait(60)
            time.sleep(10)  # wait for page to load new content
            if i % 54 == 0:
                try:
                    driver.find_element_by_xpath('/html/body/div[3]/main/div/div[1]/div/div[3]/div[2]/div[4]/div/div').click()
                    driver.implicitly_wait(20)
                    time.sleep(10)
                except:
                    print(f'\nstoped in article {i}')
                    break
        i += 1
    print(f'\n\n\nemails: {emails}\n\nto check: {to_check}\n arrived to: {i}')


